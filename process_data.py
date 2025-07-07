import mysql.connector
import pandas as pd
import json
from datetime import datetime
import csv
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv
import os
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.style import Style
from rich import box
console = Console()

# Configuración de conexión
DB_NAME = 'finanzas_rpa'
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': DB_NAME
}

# 1. Crear base de datos y tablas si no existen

def create_database_and_tables():
    conn = mysql.connector.connect(host=DB_CONFIG['host'], user=DB_CONFIG['user'], password=DB_CONFIG['password'])
    cursor = conn.cursor()
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;")
    cursor.close()
    conn.close()

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS tasas_cambio (
        id INT AUTO_INCREMENT PRIMARY KEY,
        fecha DATE NOT NULL,
        moneda VARCHAR(10) NOT NULL,
        tasa DECIMAL(15,8) NOT NULL,
        UNIQUE KEY (fecha, moneda)
    );''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS transacciones (
        id INT AUTO_INCREMENT PRIMARY KEY,
        fecha DATE NOT NULL,
        descripcion VARCHAR(255),
        tipo VARCHAR(50),
        moneda VARCHAR(10),
        monto DECIMAL(15,2),
        tasa_usd DECIMAL(15,8),
        monto_usd DECIMAL(15,2),
        tasa_id INT,
        FOREIGN KEY (tasa_id) REFERENCES tasas_cambio(id)
    );''')
    # Crear tabla de resumen financiero
    cursor.execute('''CREATE TABLE IF NOT EXISTS resumen_financiero (
        id INT AUTO_INCREMENT PRIMARY KEY,
        fecha_generacion DATETIME NOT NULL,
        balance_total_usd DECIMAL(15,2),
        total_ingresos_usd DECIMAL(15,2),
        total_gastos_usd DECIMAL(15,2),
        num_transacciones INT,
        num_ingresos INT,
        num_gastos INT
    );''')
    cursor.close()
    conn.close()

# 2. Cargar tasas de cambio desde JSON

def cargar_tasas_cambio(json_file):
    with open('data/exchange_rates.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    fecha = data['date']
    rates = data['rates']
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    for moneda, tasa in rates.items():
        cursor.execute('''INSERT IGNORE INTO tasas_cambio (fecha, moneda, tasa) VALUES (%s, %s, %s)''', (fecha, moneda, tasa))
    conn.commit()
    cursor.close()
    conn.close()

# 3. Cargar transacciones desde CSV y enlazar con tasa de cambio

def cargar_transacciones(csv_file):
    df = pd.read_csv('data/transacciones.csv')
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    for _, row in df.iterrows():
        fecha = row['Fecha']
        descripcion = row['Descripción']
        tipo = row['Tipo']
        moneda = row['Moneda']
        monto = float(row['Monto'])
        # Buscar la tasa de cambio para esa fecha y moneda
        cursor.execute('''SELECT id, tasa FROM tasas_cambio WHERE fecha=%s AND moneda=%s''', (fecha, moneda))
        result = cursor.fetchone()
        if result:
            tasa_id, tasa = result
        else:
            # Si no hay tasa para esa fecha/moneda, buscar la más reciente anterior
            cursor.execute('''SELECT id, tasa FROM tasas_cambio WHERE moneda=%s AND fecha<=%s ORDER BY fecha DESC LIMIT 1''', (moneda, fecha))
            result = cursor.fetchone()
            if result:
                tasa_id, tasa = result
            else:
                tasa_id, tasa = None, 1.0  # Default: 1.0 si no hay tasa
        monto_usd = monto / float(tasa) if tasa else monto
        cursor.execute('''INSERT INTO transacciones (fecha, descripcion, tipo, moneda, monto, tasa_usd, monto_usd, tasa_id)
                          VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                       (fecha, descripcion, tipo, moneda, monto, tasa, monto_usd, tasa_id))
    conn.commit()
    cursor.close()
    conn.close()

# 4. Consulta ejemplo: transacciones con tasa y monto en USD

def consultar_transacciones():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute('''SELECT t.id, t.fecha, t.descripcion, t.tipo, t.moneda, t.monto, tc.tasa, t.monto_usd
                      FROM transacciones t
                      LEFT JOIN tasas_cambio tc ON t.tasa_id = tc.id
                      ORDER BY t.fecha, t.id''')
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results

# 5. Main: orquestar el flujo

def calcular_y_guardar_resumen():
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    # Balance total
    cursor.execute('SELECT SUM(monto_usd) as balance_total FROM transacciones')
    balance_total = cursor.fetchone()['balance_total'] or 0
    # Total ingresos
    cursor.execute("SELECT SUM(monto_usd) as total_ingresos FROM transacciones WHERE tipo='Ingreso'")
    total_ingresos = cursor.fetchone()['total_ingresos'] or 0
    # Total gastos
    cursor.execute("SELECT SUM(monto_usd) as total_gastos FROM transacciones WHERE tipo='Gasto'")
    total_gastos = cursor.fetchone()['total_gastos'] or 0
    # Número de transacciones
    cursor.execute('SELECT COUNT(*) as num_transacciones FROM transacciones')
    num_transacciones = cursor.fetchone()['num_transacciones'] or 0
    # Número de ingresos
    cursor.execute("SELECT COUNT(*) as num_ingresos FROM transacciones WHERE tipo='Ingreso'")
    num_ingresos = cursor.fetchone()['num_ingresos'] or 0
    # Número de gastos
    cursor.execute("SELECT COUNT(*) as num_gastos FROM transacciones WHERE tipo='Gasto'")
    num_gastos = cursor.fetchone()['num_gastos'] or 0
    # Guardar en la tabla resumen_financiero
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor2 = conn.cursor()
    cursor2.execute('''INSERT INTO resumen_financiero (fecha_generacion, balance_total_usd, total_ingresos_usd, total_gastos_usd, num_transacciones, num_ingresos, num_gastos)
                       VALUES (%s, %s, %s, %s, %s, %s, %s)''',
                    (now, balance_total, total_ingresos, total_gastos, num_transacciones, num_ingresos, num_gastos))
    conn.commit()
    cursor2.close()
    cursor.close()
    conn.close()
    # Imprimir resumen financiero con Rich
    table = Table(title="Resumen Financiero", box=box.ROUNDED, style=Style(color="white", bgcolor="#2a5298"))
    table.add_column("Campo", style="bold white", header_style="bold white on #2a5298")
    table.add_column("Valor", style="bold white", header_style="bold white on #2a5298")
    table.add_row("Fecha de generación", now)
    table.add_row("Balance total en USD", f"{balance_total:.2f}")
    table.add_row("Total Ingresos en USD", f"{total_ingresos:.2f}")
    table.add_row("Total Gastos en USD", f"{total_gastos:.2f}")
    table.add_row("Número de transacciones", str(num_transacciones))
    table.add_row("Ingresos", str(num_ingresos))
    table.add_row("Gastos", str(num_gastos))
    console.print(table)

    # Exportar a JSON
    resumen = {
        'fecha_generacion': now,
        'balance_total_usd': float(balance_total),
        'total_ingresos_usd': float(total_ingresos),
        'total_gastos_usd': float(total_gastos),
        'num_transacciones': int(num_transacciones),
        'num_ingresos': int(num_ingresos),
        'num_gastos': int(num_gastos)
    }
    with open('reportes/resumen_financiero.json', 'w', encoding='utf-8') as fjson:
        json.dump(resumen, fjson, ensure_ascii=False, indent=2)
    # Exportar a XML
    root = ET.Element('ResumenFinanciero')
    for k, v in resumen.items():
        child = ET.SubElement(root, k)
        child.text = str(v)
    tree = ET.ElementTree(root)
    tree.write('reportes/resumen_financiero.xml', encoding='utf-8', xml_declaration=True)
    # Definir encabezados y valores para la hoja de resumen ejecutivo
    encabezados = [
        'Fecha de generación',
        'Balance Total (USD)',
        'Total Ingresos (USD)',
        'Total Gastos (USD)',
        'N° Transacciones',
        'N° Ingresos',
        'N° Gastos'
    ]
    valores = [
        resumen['fecha_generacion'],
        resumen['balance_total_usd'],
        resumen['total_ingresos_usd'],
        resumen['total_gastos_usd'],
        resumen['num_transacciones'],
        resumen['num_ingresos'],
        resumen['num_gastos']
    ]
    # Exportar a Excel (.xlsx) con varias hojas profesionales
    wb = Workbook()
    # --- Hoja 1: Resumen Ejecutivo ---
    ws1 = wb.active
    ws1.title = 'Resumen Ejecutivo'
    for col, encabezado in enumerate(encabezados, 1):
        cell = ws1.cell(row=1, column=col, value=encabezado)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, valor in enumerate(valores, 1):
        cell = ws1.cell(row=2, column=col, value=valor)
        cell.font = Font(bold=False, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws1.column_dimensions[column].width = max_length + 2
    ws1['A2'].number_format = 'DD/MM/YYYY HH:MM:SS'
    for col in ['B', 'C', 'D']:
        ws1[f'{col}2'].number_format = '#,##0.00'

    # --- Hoja 2: Transacciones (solo las actuales) ---
    ws2 = wb.create_sheet('Transacciones')
    trans_headers = ['ID', 'Fecha', 'Descripción', 'Tipo', 'Moneda', 'Monto', 'Tasa USD', 'Monto USD']
    with mysql.connector.connect(**DB_CONFIG) as conn2:
        with conn2.cursor() as cursor:
            for col, encabezado in enumerate(trans_headers, 1):
                cell = ws2.cell(row=1, column=col, value=encabezado)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cursor.execute('SELECT id, fecha, descripcion, tipo, moneda, monto, tasa_usd, monto_usd FROM transacciones ORDER BY fecha, id')
            trans_rows = cursor.fetchall()
            for row_idx, row in enumerate(trans_rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col in ws2.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws2.column_dimensions[column].width = max_length + 2

    # --- Hoja 3: Tasas de Cambio (solo las actuales) ---
    ws3 = wb.create_sheet('Tasas de Cambio')
    tasas_headers = ['ID', 'Fecha', 'Moneda', 'Tasa']
    with mysql.connector.connect(**DB_CONFIG) as conn3:
        with conn3.cursor() as cursor:
            for col, encabezado in enumerate(tasas_headers, 1):
                cell = ws3.cell(row=1, column=col, value=encabezado)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cursor.execute('SELECT id, fecha, moneda, tasa FROM tasas_cambio ORDER BY fecha, moneda')
            tasas_rows = cursor.fetchall()
            for row_idx, row in enumerate(tasas_rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col in ws3.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws3.column_dimensions[column].width = max_length + 2

    # --- Hoja 4: Historial (todas las transacciones y tasas) ---
    ws4 = wb.create_sheet('Historial')
    hist_headers = ['ID', 'Fecha', 'Descripción', 'Tipo', 'Moneda', 'Monto', 'Tasa USD', 'Monto USD']
    with mysql.connector.connect(**DB_CONFIG) as conn4:
        with conn4.cursor() as cursor:
            for col, encabezado in enumerate(hist_headers, 1):
                cell = ws4.cell(row=1, column=col, value=encabezado)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cursor.execute('SELECT id, fecha, descripcion, tipo, moneda, monto, tasa_usd, monto_usd FROM transacciones ORDER BY fecha, id')
            hist_rows = cursor.fetchall()
            for row_idx, row in enumerate(hist_rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col in ws4.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws4.column_dimensions[column].width = max_length + 2

    wb.save('reportes/resumen_financiero.xlsx')
    enviar_resumen_por_email(resumen)

def enviar_resumen_por_email(resumen):
    load_dotenv()
    remitente = os.getenv('EMAIL_USER')
    password = os.getenv('EMAIL_PASS')
    
    # Obtener emails desde la base de datos
    from email_manager import get_email_list
    destinatario = get_email_list()
    
    if not destinatario:
        print("⚠️ No hay emails configurados en la base de datos")
        return

    print(f"Intentando enviar email a: {destinatario}")
    print(f"Desde: {remitente}")

    asunto = "Reporte Ejecutivo: Resumen Financiero"
    # Cuerpo HTML elegante
    cuerpo_html = f'''
    <html>
    <body style="font-family: Arial, sans-serif; background-color: #f4f6f8; color: #222;">
        <div style="max-width: 600px; margin: auto; background: #fff; border-radius: 10px; box-shadow: 0 2px 8px #ccc; padding: 32px;">
            <h2 style="color: #2a5298; text-align: center;">Reporte Ejecutivo: Resumen Financiero</h2>
            <p>Estimado equipo,</p>
            <p>Adjunto encontrarán el reporte ejecutivo generado automáticamente por el sistema de RPA.</p>
            <table style="width: 100%; border-collapse: collapse; margin: 24px 0;">
                <tr style="background-color: #2a5298; color: #fff;">
                    <th style="padding: 8px;">Fecha de generación</th>
                    <th style="padding: 8px;">Balance Total (USD)</th>
                    <th style="padding: 8px;">Total Ingresos (USD)</th>
                    <th style="padding: 8px;">Total Gastos (USD)</th>
                    <th style="padding: 8px;">N° Transacciones</th>
                    <th style="padding: 8px;">N° Ingresos</th>
                    <th style="padding: 8px;">N° Gastos</th>
                </tr>
                <tr style="background-color: #eaf0fa;">
                    <td style="padding: 8px; text-align: center;">{resumen['fecha_generacion']}</td>
                    <td style="padding: 8px; text-align: right;">{resumen['balance_total_usd']:.2f}</td>
                    <td style="padding: 8px; text-align: right;">{resumen['total_ingresos_usd']:.2f}</td>
                    <td style="padding: 8px; text-align: right;">{resumen['total_gastos_usd']:.2f}</td>
                    <td style="padding: 8px; text-align: center;">{resumen['num_transacciones']}</td>
                    <td style="padding: 8px; text-align: center;">{resumen['num_ingresos']}</td>
                    <td style="padding: 8px; text-align: center;">{resumen['num_gastos']}</td>
                </tr>
            </table>
            <p style="margin-top: 24px;">El archivo Excel adjunto contiene:</p>
            <ul>
                <li>Resumen ejecutivo</li>
                <li>Transacciones</li>
                <li>Tasas de cambio</li>
                <li>Historial consolidado</li>
            </ul>
            <p style="margin-top: 32px; color: #555;">Cualquier duda o comentario, quedo atento.<br><br>Saludos cordiales,<br><b>Automatización RPA Luis Viña</b></p>
        </div>
    </body>
    </html>
    '''
    
    try:
        msg = EmailMessage()
        msg['Subject'] = asunto
        msg['From'] = remitente
        msg['To'] = destinatario
        msg.set_content("Este correo contiene un resumen financiero en formato HTML. Si no lo visualiza correctamente, por favor revise el adjunto.")
        msg.add_alternative(cuerpo_html, subtype='html')

        # Verificar que existe el archivo antes de adjuntarlo
        if os.path.exists('reportes/resumen_financiero.xlsx'):
            with open('reportes/resumen_financiero.xlsx', 'rb') as f:
                file_data = f.read()
                file_name = 'resumen_financiero.xlsx'
            msg.add_attachment(
                file_data,
                maintype='application',
                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=file_name
            )
            print("Archivo Excel adjuntado correctamente")
        else:
            print("Archivo Excel no encontrado, enviando sin adjunto")

        print("Conectando a Gmail...")
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            print("Autenticando...")
            smtp.login(remitente, password)
            print("Enviando email...")
            smtp.send_message(msg)
            print("Email enviado exitosamente")
            
        console.print(Panel(f"[OK] Correo enviado exitosamente a {destinatario} con el resumen financiero adjunto.", style=Style(color="white", bgcolor="#28a745"), box=box.ROUNDED))
        
    except smtplib.SMTPAuthenticationError as e:
        error_msg = f"Error de autenticación: Verifica EMAIL_USER y EMAIL_PASS en .env"
        print(error_msg)
        console.print(Panel(error_msg, style=Style(color="white", bgcolor="#dc3545"), box=box.ROUNDED))
        
    except smtplib.SMTPRecipientsRefused as e:
        error_msg = f"Error con destinatarios: {e}"
        print(error_msg)
        console.print(Panel(error_msg, style=Style(color="white", bgcolor="#dc3545"), box=box.ROUNDED))
        
    except smtplib.SMTPServerDisconnected as e:
        error_msg = f"Error de conexión con servidor: {e}"
        print(error_msg)
        console.print(Panel(error_msg, style=Style(color="white", bgcolor="#dc3545"), box=box.ROUNDED))
        
    except Exception as e:
        error_msg = f"Error inesperado enviando email: {e}"
        print(error_msg)
        console.print(Panel(error_msg, style=Style(color="white", bgcolor="#dc3545"), box=box.ROUNDED))

def main():
    create_database_and_tables()
    cargar_tasas_cambio('exchange_rates.json')
    cargar_transacciones('transacciones.csv')
    calcular_y_guardar_resumen()

if __name__ == '__main__':
    main() 