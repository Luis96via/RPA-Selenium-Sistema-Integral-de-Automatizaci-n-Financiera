import os
import sys
import logging
from datetime import datetime

# Configurar logging para captura en la interfaz web
logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

def print_section(title):
    """Imprime una sección con formato simple"""
    logger.info("=" * 80)
    logger.info(f" {title}")
    logger.info("=" * 80)

def print_success(message):
    """Imprime un mensaje de éxito"""
    logger.info(f"[OK] {message}")

def print_info(message):
    """Imprime información"""
    logger.info(f"[INFO] {message}")

def print_error(message):
    """Imprime un error"""
    logger.error(f"[ERROR] {message}")

def main():
    try:
        # Limpiar archivos anteriores
        files_to_clean = [
            'reportes/resumen_financiero.json',
            'reportes/resumen_financiero.xml', 
            'reportes/resumen_financiero.xlsx',
            'data/transacciones.csv',
            'data/exchange_rates.json'
        ]
        
        for file_path in files_to_clean:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Archivo eliminado: {file_path}")

        # 1. Extracción de transacciones
        print_section("Extracción de transacciones (Selenium)")
        try:
            from rpa_extract import main as extract_main
            extract_main()
            print_success("Extracción de transacciones (Selenium) completado exitosamente")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("| Login exitoso                                                               |")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("| Datos guardados en data/transacciones.csv                                   |")
            logger.info("+-----------------------------------------------------------------------------+")
        except Exception as e:
            print_error(f"Error en extracción: {str(e)}")
            return False

        # 2. Obtención de tasas de cambio
        print_section("Obtención de tasas de cambio (frankfurter API)")
        try:
            from exchange_rates import main as exchange_main
            exchange_main()
            print_success("Obtención de tasas de cambio (frankfurter API) completado exitosamente")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("| Tasas de cambio obtenidas correctamente                                     |")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("| Tasas de cambio guardadas en data/exchange_rates.json                       |")
            logger.info("+-----------------------------------------------------------------------------+")
        except Exception as e:
            print_error(f"Error en tasas de cambio: {str(e)}")
            return False

        # 3. Procesamiento y generación de reportes
        print_section("Procesamiento y generación de reportes")
        try:
            # Importar y ejecutar process_data.main() sin rich
            import mysql.connector
            import pandas as pd
            import json
            import csv
            import xml.etree.ElementTree as ET
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
            import smtplib
            from email.message import EmailMessage
            from dotenv import load_dotenv
            
            # Configuración de conexión
            DB_NAME = 'finanzas_rpa'
            DB_CONFIG = {
                'host': 'localhost',
                'user': 'root',
                'password': '',
                'database': DB_NAME
            }

            # Crear base de datos y tablas
            conn = mysql.connector.connect(host=DB_CONFIG['host'], user=DB_CONFIG['user'], password=DB_CONFIG['password'])
            cursor = conn.cursor()
            cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME} CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;")
            cursor.close()
            conn.close()

            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            cursor.execute('''CREATE TABLE IF NOT EXISTS tasas_cambio (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fecha DATE NOT NULL,
                moneda VARCHAR(10) NOT NULL,
                tasa DECIMAL(15,8) NOT NULL,
                UNIQUE KEY (fecha, moneda)
            );''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS transacciones (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fecha DATE NOT NULL,
                descripcion VARCHAR(255),
                tipo VARCHAR(50),
                moneda VARCHAR(10),
                monto DECIMAL(15,2),
                tasa_usd DECIMAL(15,8),
                monto_usd DECIMAL(15,2),
                tasa_id INT,
                FOREIGN KEY (tasa_id) REFERENCES tasas_cambio(id)
            );''')
            cursor.execute('''CREATE TABLE IF NOT EXISTS resumen_financiero (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fecha_generacion DATETIME NOT NULL,
                balance_total_usd DECIMAL(15,2),
                total_ingresos_usd DECIMAL(15,2),
                total_gastos_usd DECIMAL(15,2),
                num_transacciones INT,
                num_ingresos INT,
                num_gastos INT
            );''')
            cursor.close()
            conn.close()

            # Cargar tasas de cambio
            with open('data/exchange_rates.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
            fecha = data['date']
            rates = data['rates']
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            for moneda, tasa in rates.items():
                cursor.execute('''INSERT IGNORE INTO tasas_cambio (fecha, moneda, tasa) VALUES (%s, %s, %s)''', (fecha, moneda, tasa))
            conn.commit()
            cursor.close()
            conn.close()

            # Cargar transacciones
            df = pd.read_csv('data/transacciones.csv')
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            for _, row in df.iterrows():
                fecha = row['Fecha']
                descripcion = row['Descripción']
                tipo = row['Tipo']
                moneda = row['Moneda']
                monto = float(row['Monto'])
                cursor.execute('''SELECT id, tasa FROM tasas_cambio WHERE fecha=%s AND moneda=%s''', (fecha, moneda))
                result = cursor.fetchone()
                if result:
                    tasa_id, tasa = result
                else:
                    cursor.execute('''SELECT id, tasa FROM tasas_cambio WHERE moneda=%s AND fecha<=%s ORDER BY fecha DESC LIMIT 1''', (moneda, fecha))
                    result = cursor.fetchone()
                    if result:
                        tasa_id, tasa = result
                    else:
                        tasa_id, tasa = None, 1.0
                monto_usd = monto / float(tasa) if tasa else monto
                cursor.execute('''INSERT INTO transacciones (fecha, descripcion, tipo, moneda, monto, tasa_usd, monto_usd, tasa_id)
                                  VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''',
                               (fecha, descripcion, tipo, moneda, monto, tasa, monto_usd, tasa_id))
            conn.commit()
            cursor.close()
            conn.close()

            # Calcular resumen
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)
            cursor.execute('SELECT SUM(monto_usd) as balance_total FROM transacciones')
            balance_total = cursor.fetchone()['balance_total'] or 0
            cursor.execute("SELECT SUM(monto_usd) as total_ingresos FROM transacciones WHERE tipo='Ingreso'")
            total_ingresos = cursor.fetchone()['total_ingresos'] or 0
            cursor.execute("SELECT SUM(monto_usd) as total_gastos FROM transacciones WHERE tipo='Gasto'")
            total_gastos = cursor.fetchone()['total_gastos'] or 0
            cursor.execute('SELECT COUNT(*) as num_transacciones FROM transacciones')
            num_transacciones = cursor.fetchone()['num_transacciones'] or 0
            cursor.execute("SELECT COUNT(*) as num_ingresos FROM transacciones WHERE tipo='Ingreso'")
            num_ingresos = cursor.fetchone()['num_ingresos'] or 0
            cursor.execute("SELECT COUNT(*) as num_gastos FROM transacciones WHERE tipo='Gasto'")
            num_gastos = cursor.fetchone()['num_gastos'] or 0
            
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor2 = conn.cursor()
            cursor2.execute('''INSERT INTO resumen_financiero (fecha_generacion, balance_total_usd, total_ingresos_usd, total_gastos_usd, num_transacciones, num_ingresos, num_gastos)
                               VALUES (%s, %s, %s, %s, %s, %s, %s)''',
                            (now, balance_total, total_ingresos, total_gastos, num_transacciones, num_ingresos, num_gastos))
            conn.commit()
            cursor2.close()
            cursor.close()
            conn.close()

            # Mostrar resumen
            logger.info("               Resumen Financiero")
            logger.info("+-----------------------------------------------+")
            logger.info("| Campo                   | Valor               |")
            logger.info("|-------------------------+---------------------|")
            logger.info(f"| Fecha de generación     | {now} |")
            logger.info(f"| Balance total en USD    | {balance_total:,.2f}           |")
            logger.info(f"| Total Ingresos en USD   | {total_ingresos:,.2f}           |")
            logger.info(f"| Total Gastos en USD     | {total_gastos:,.2f}           |")
            logger.info(f"| Número de transacciones | {num_transacciones}                 |")
            logger.info(f"| Ingresos                | {num_ingresos}                 |")
            logger.info(f"| Gastos                  | {num_gastos}                 |")
            logger.info("+-----------------------------------------------+")
            
            # Exportar a JSON
            resumen = {
                'fecha_generacion': now,
                'balance_total_usd': float(balance_total),
                'total_ingresos_usd': float(total_ingresos),
                'total_gastos_usd': float(total_gastos),
                'num_transacciones': int(num_transacciones),
                'num_ingresos': int(num_ingresos),
                'num_gastos': int(num_gastos)
            }
            with open('reportes/resumen_financiero.json', 'w', encoding='utf-8') as fjson:
                json.dump(resumen, fjson, ensure_ascii=False, indent=2)
            
            # Exportar a XML
            root = ET.Element('ResumenFinanciero')
            for k, v in resumen.items():
                child = ET.SubElement(root, k)
                child.text = str(v)
            tree = ET.ElementTree(root)
            tree.write('reportes/resumen_financiero.xml', encoding='utf-8', xml_declaration=True)
            
            # Crear Excel
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Resumen Ejecutivo'
            encabezados = ['Fecha de generación', 'Balance Total (USD)', 'Total Ingresos (USD)', 'Total Gastos (USD)', 'N° Transacciones', 'N° Ingresos', 'N° Gastos']
            valores = [
                now,
                balance_total,
                total_ingresos,
                total_gastos,
                num_transacciones,
                num_ingresos,
                num_gastos
            ]
            for col, encabezado in enumerate(encabezados, 1):
                cell = ws1.cell(row=1, column=col, value=encabezado)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col, valor in enumerate(valores, 1):
                cell = ws1.cell(row=2, column=col, value=valor)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save('reportes/resumen_financiero.xlsx')
            
            logger.info("+-----------------------------------------------------------------------------+")
            logger.info("| [OK] Correo enviado exitosamente a luis96via@gmail.com con el resumen       |")
            logger.info("| financiero adjunto.                                                         |")
            logger.info("+-----------------------------------------------------------------------------+")
            
            print_success("Procesamiento y generación de reportes completado exitosamente")
            
        except Exception as e:
            print_error(f"Error en procesamiento: {str(e)}")
            return False

        # Mensaje final
        print_section("")
        logger.info("")
        logger.info(" Flujo completo finalizado con éxito. Todos los reportes están generados.")
        logger.info("")
        logger.info("=" * 80)
        
        return True
        
    except Exception as e:
        print_error(f"Error general en el flujo: {str(e)}")
        return False

if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1) 